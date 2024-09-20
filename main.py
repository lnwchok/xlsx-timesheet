from openpyxl import Workbook, load_workbook

filename = r"C:\Users\QX6-XXF1808D002\Desktop\Doc List.xlsx"

wb = load_workbook(filename, read_only=True, keep_vba=False)
ws = wb["MESA Doc List to EVLS"]

data = ws["B6"].value

print(data)
